from telegram.ext import Updater
from telegram.ext import CommandHandler
from telegram.ext import MessageHandler, Filters
import logging
import requests
from os.path import abspath, join, dirname, exists


class Config(object):
    def __init__(self):
        from os import chdir
        from os.path import abspath
        import json

        chdir(dirname(abspath(__file__)))
        with open('config.cfg', 'r') as config_file:
            j = json.load(config_file)
        self.__dict__ = j


# - Config loading
config = Config()

botState = 'idle'
botStates = ('idle', 'waitingForCounter44', 'waitingForCounter13', 'waitingForCounter26',
    'waitingForCounter53', 'waitingForCounter_el_suv472', 'waitingForCounter82',
    'waitingForCounter31', 'waitingForCounter_el_suv662')
updater = Updater(token=config.telegram_updater_token)
dispatcher = updater.dispatcher
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

counter44 = 0
counter13 = 0
counter26 = 0
counter53 = 0
counter_el_suv472 = 0
counter82 = 0
counter31 = 0
counter_el_suv662 = 0

previousPokazaniyaDate = ''
prevCounter44 = 0
prevCounter13 = 0
prevCounter26 = 0
prevCounter53 = 0
prevCounter_el_suv472 = 0
prevCounter82 = 0
prevCounter31 = 0
prevCounter_el_suv662 = 0
previousDict = {}

# excelFilename = 'C:\\Users\\kvant\\YandexDisk\\Квартира_Родственники\\Показания.xlsx'
excelFilename = join(dirname(abspath(__file__)), 'Показания.xlsx')
chat_id = ''

# - yandex.api
disk_xls_path = 'Квартира_Родственники/Показания.xlsx'
disk_token = config.disk_token
yandex_host = 'https://cloud-api.yandex.net'


def ResetBot():
    global botState
    global counter44, counter13, counter26, counter53, counter_el_suv472, counter82, counter31, counter_el_suv662
    global previousPokazaniyaDate, prevCounter44, prevCounter13, prevCounter26, prevCounter53, prevCounter_el_suv472, prevCounter82, prevCounter31, prevCounter_el_suv662
    global previousDict
    botState = 'idle'
    counter44 = 0
    counter13 = 0
    counter26 = 0
    counter53 = 0
    counter_el_suv472 = 0
    counter82 = 0
    counter31 = 0
    counter_el_suv662 = 0
    previousPokazaniyaDate = ''
    prevCounter44 = 0
    prevCounter13 = 0
    prevCounter26 = 0
    prevCounter53 = 0
    prevCounter_el_suv472 = 0
    prevCounter82 = 0
    prevCounter31 = 0
    prevCounter_el_suv662 = 0
    previousDict.clear()


def GetFileFromYandexDisk(save_to=disk_xls_path, excelFilename=excelFilename):
    '''Yandex.disk API file download'''

    global yandex_host, ses, chat_id

    yandex_dl_method = yandex_host + '/v1/disk/resources/download'
    yandex_dl_params = {'path': disk_xls_path}

    j = ses.get(yandex_dl_method, params=yandex_dl_params).json()
    if 'error' in j:  # throw exception
        print(j['error'])
        print(j['message'])
        print(j['description'])
        updater.bot.send_message(chat_id=chat_id, text=j['error'] + j['message'] + j['description'])
    else:
        method = ses.__getattribute__(j['method'].lower())

        with open(excelFilename, 'wb') as xls_file:
            r = method(j['href'])
            for chunk in r.iter_content(1024):
                xls_file.write(chunk)

    return exists(excelFilename)


def UploadFileToYandexDisk(excelFilename=excelFilename):
    '''Yandex.disk API file upload'''
    from time import sleep
    from os import remove

    yandex_upload_method = yandex_host + '/v1/disk/resources/upload'
    yandex_upload_params = {'path': disk_xls_path, 'overwrite': 'true'}
    j = ses.get(yandex_upload_method, params=yandex_upload_params).json()

    # - Отправляем файл в Яндекс.Диск
    with open(excelFilename, 'rb') as xls_file:
        p = ses.put(j['href'], data=xls_file)
    if p.status_code != 201:
        print(f'Failed file uploading ({p.status_code})')

    # - Check operation status ('in-progress', 'success')
    operation_id = j['operation_id']
    while True:
        j = ses.get(yandex_host + '/v1/disk/operations/' + operation_id).json()
        if 'status' not in j:
            print('Error in getting operation status')
            print(j)
            break

        if j['status'].lower() != 'success':
            print('Warning: Operation still not success. Waiting 1 sec...')
            sleep(1)
        else:
            print('Good')
            if exists(excelFilename):
                remove(excelFilename)
            break


def ReadFromXL(xlFilename=excelFilename):
    global counter44, counter13, counter26, counter53, counter_el_suv472, counter82, counter31, counter_el_suv662
    global previousPokazaniyaDate, prevCounter44, prevCounter13, prevCounter26, prevCounter53, prevCounter_el_suv472, prevCounter82, prevCounter31, prevCounter_el_suv662
    import openpyxl
    import datetime

    wb = openpyxl.load_workbook(xlFilename)
    wb.get_sheet_names()
    sheet = wb.get_sheet_by_name('Лист1')

    row = sheet.max_row - 1
    previousPokazaniyaDate = sheet[f'A{row}'].value
    prevCounter44 = sheet[f'B{row}'].value
    prevCounter13 = sheet[f'C{row}'].value
    prevCounter26 = sheet[f'D{row}'].value
    prevCounter53 = sheet[f'E{row}'].value
    prevCounter_el_suv472 = sheet[f'F{row}'].value
    prevCounter82 = sheet[f'G{row}'].value
    prevCounter31 = sheet[f'H{row}'].value
    prevCounter_el_suv662 = sheet[f'I{row}'].value
    prevDict = {}
    prevDict['44'] = prevCounter44
    prevDict['13'] = prevCounter13
    prevDict['26'] = prevCounter26
    prevDict['53'] = prevCounter53
    prevDict['el_suv472'] = prevCounter_el_suv472
    prevDict['82'] = prevCounter82
    prevDict['31'] = prevCounter31
    prevDict['el_suv662'] = prevCounter_el_suv662
    return prevDict


def writeToXL(xlFilename=excelFilename):
    global counter44, counter13, counter26, counter53, counter_el_suv472, counter82, counter31, counter_el_suv662
    import openpyxl
    import copy
    import datetime

    wb = openpyxl.load_workbook(xlFilename)
    wb.get_sheet_names()
    sheet = wb.get_sheet_by_name('Лист1')

    maxCol = sheet.max_column
    maxRow = sheet.max_row
    openpyxl.worksheet.worksheet.Worksheet.insert_rows(sheet, idx=maxRow, amount=1)

    for i_col in range(1, maxCol + 1, 1):
        source_cell = sheet.cell(row=maxRow - 1, column=i_col)
        target_cell = sheet.cell(row=maxRow, column=i_col)
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
    sheet[f'A{maxRow}'] = datetime.datetime.now()
    sheet[f'B{maxRow}'] = counter44  # 44
    sheet[f'C{maxRow}'] = counter13  # 13
    sheet[f'D{maxRow}'] = counter26  # 26
    sheet[f'E{maxRow}'] = counter53  # 53
    sheet[f'F{maxRow}'] = counter_el_suv472  # electro
    sheet[f'G{maxRow}'] = counter82  # 82
    sheet[f'H{maxRow}'] = counter31  # 31
    sheet[f'I{maxRow}'] = counter_el_suv662  # electro

    wb.save(xlFilename)


def testForState():
    global botState
    return botState != 'waitingForcounter44'


def start(update, context):
    bot = context.bot
    s = 'Привет! Я бот, который позволит быстро создать сообщения для отправки коммунальным службам,\
а также вести архив и отчёт в Excel-файле на твоём Яндекс.Диске.'
    bot.sendMessage(chat_id=update.message.chat_id, text=s)


def sendCounters(update, context):
    global botState, previousDict
    global chat_id
    bot = context.bot
    testForState()
    previousDict = ReadFromXL()
    s = f"Введите показания счётчика 44 (прошлые показания: {previousDict['44']}) :"
    botState = 'waitingForCounter44'
    chat_id = update.message.chat_id
    print(chat_id)
    bot.sendMessage(chat_id=update.message.chat_id, text=s)


def getPreviousPokazaniya(update, context):
    bot = context.bot
    previous_pokazaniya = ReadFromXL()
    s = ''
    for schetchik in previous_pokazaniya:
        s += f'{schetchik}: {previous_pokazaniya[schetchik]}\n'
    bot.sendMessage(chat_id=update.message.chat_id, text=s)


# - если пришло любое сообщение, не команда
#def echo(bot, update):
def echo(update, context):
    global botState
    global counter44, counter13, counter26, counter53, counter_el_suv472, counter82, counter31, counter_el_suv662
    global excelFilename
    global previousDict
    bot = context.bot
    testForState()
    try:
        if update.message.text.strip() == '-':
            ResetBot()
            bot.sendMessage(chat_id=update.message.chat_id, text=f'Сброс режимов. ({counter44})')
            return

        elif botState == 'waitingForCounter44':
            counter44 = int(update.message.text)
            bot.sendMessage(chat_id=update.message.chat_id, text=f"Введите показания счётчика 13  (прошлые показания: {previousDict['13']}) :")
            botState = 'waitingForCounter13'
        elif botState == 'waitingForCounter13':
            counter13 = int(update.message.text)
            bot.sendMessage(chat_id=update.message.chat_id, text=f"Введите показания счётчика 26 (прошлые показания: {previousDict['26']}) :")
            botState = 'waitingForCounter26'
        elif botState == 'waitingForCounter26':
            counter26 = int(update.message.text)
            bot.sendMessage(chat_id=update.message.chat_id, text=f"Введите показания счётчика 53 (прошлые показания: {previousDict['53']}) :")
            botState = 'waitingForCounter53'
        elif botState == 'waitingForCounter53':
            counter53 = int(update.message.text)
            bot.sendMessage(chat_id=update.message.chat_id, text=f"Введите показания счётчика Электросчётчик Суворова 4-72 (прошлые показания: {previousDict['el_suv472']}) :")
            botState = 'waitingForCounter_el_suv472'
        elif botState == 'waitingForCounter_el_suv472':
            counter_el_suv472 = int(update.message.text)
            bot.sendMessage(chat_id=update.message.chat_id, text=f"Введите показания счётчика 82 (прошлые показания: {previousDict['82']}) :")
            botState = 'waitingForCounter82'
        elif botState == 'waitingForCounter82':
            counter82 = int(update.message.text)
            bot.sendMessage(chat_id=update.message.chat_id, text=f"Введите показания счётчика 31 (прошлые показания: {previousDict['31']}) :")
            botState = 'waitingForCounter31'
        elif botState == 'waitingForCounter31':
            counter31 = int(update.message.text)
            bot.sendMessage(chat_id=update.message.chat_id, text=f"Введите показания счётчика Электросчётчик Суворова 6-62 (прошлые показания: {previousDict['el_suv662']}) :")
            botState = 'waitingForCounter_el_suv662'
        elif botState == 'waitingForCounter_el_suv662':
            counter_el_suv662 = int(update.message.text)
            s = f'Показания приняты.\n1575544 = {counter44}\n1579913 = {counter13}\
            \n1579926 = {counter26}\n29753953 = {counter53}\nЭлектросчётчик Основной = {counter_el_suv472}\
            \n25601382 = {counter82}\n11200331 = {counter31}\nЭлектросчётчик Доп. = {counter_el_suv662}'
            botState = 'CountersAccepted'

            # - Заносим в Excel
            writeToXL()

            # - Формируем сообщения в ВОТСАП
            # --- Энергосбыт:
            waEnergosbyt = f'''`Лиц.счёт: {config.personal_account_main_energosbyt}
Горячая вода:
счётчик "1575544" - {counter44}
счётчик "1579913" - {counter13}
электросчётчик - {counter_el_suv472}

Лиц.счёт: {config.personal_account_dop_energosbyt}
Горячая вода:
счётчик "25601382" - {counter82}
электросчётчик - {counter_el_suv662}`'''

            # --- Водоканал:
            waVodokanal = f'''`{config.personal_account_main_vodokanal}
1575544-{counter44}
1579913-{counter13}
1579926-{counter26}
29753953-{counter53}
{config.personal_account_dop_vodokanal}
25601382-{counter82}
11200331-{counter31}`'''
            
            # - Concatenate two templates
            s += '\n\nСообщение для *Энергосбыта*:\n' + waEnergosbyt + '\n\nСообщение для *Водоканала*\
:\n' + waVodokanal + '\n\n'

            # - Send msg to user, обратим внимание на parse_mode == 'Markdown' для форматирования
            bot.sendMessage(chat_id=update.message.chat_id, text=s, parse_mode='Markdown')

            # - Сбрасываем состояние телеграм-бота
            ResetBot()

            # - Загружаем файл на Яндекс.Диск и удаляем файл
            UploadFileToYandexDisk()

    except Exception as e:
        bot.sendMessage(chat_id=update.message.chat_id, text='Некорректный ввод')
        print(e)
        # print(e.message)


################################################################################

# - Yandex.Disk
headers = {'Accept': 'application/json',
           'Authorization': 'OAuth ' + disk_token}
ses = requests.Session()
ses.headers.update(headers)
if not GetFileFromYandexDisk():
    print('File not downloaded')

# - обработчик команд
start_handler = CommandHandler('start', start)
sendCounters_handler = CommandHandler('sendCounters', sendCounters)  # sendCounters
getPokazaniya_handler = CommandHandler('getPokazaniya', getPreviousPokazaniya)  # getPokazaniya
dispatcher.add_handler(start_handler)
dispatcher.add_handler(sendCounters_handler)
dispatcher.add_handler(getPokazaniya_handler)

# - обработчик любых сообщений, не команд
echo_handler = MessageHandler(Filters.text, echo)
dispatcher.add_handler(echo_handler)

updater.start_polling()
print('бот запущен')

